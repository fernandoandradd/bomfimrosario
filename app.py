import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

st.set_page_config(
    page_title="Gestão de Estoque — Físico x Fiscal",
    page_icon="📦",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════════════════════
# MAPEAMENTOS DE COLUNAS
# ══════════════════════════════════════════════════════════════════════════════

COL_MAP_FISICO = {
    "Código": 0, "Produto": 1, "UN": 9, "Qtde": 10,
    "Preço Custo": 12, "Custo Final": 14, "Valor Total": 16,
}
COL_MAP_FISCAL = {
    "Código": 0, "NCM": 1, "Produto": 2, "UN": 10,
    "Último Custo": 12, "Qtde": 13, "Custo Médio": 14, "Valor Total": 16,
}

SKIP_KEYWORDS = {"Código", "Codigo", "Software C-Plus", "Registro de inventário", "Page "}


# ══════════════════════════════════════════════════════════════════════════════
# FUNÇÕES DE PROCESSAMENTO
# ══════════════════════════════════════════════════════════════════════════════

def is_data_row(row: tuple) -> bool:
    codigo = row[0]
    if codigo is None:
        return False
    if hasattr(codigo, "strftime"):
        return False
    for kw in SKIP_KEYWORDS:
        if kw in str(codigo):
            return False
    for cell in row:
        if cell is not None and isinstance(cell, str):
            s = cell.strip()
            if s.startswith("Page ") or "Todas" in s or "Registro de" in s or "Software C-Plus" in s:
                return False
    return bool(str(codigo).strip())


def detect_layout(file) -> str:
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(max_row=15, values_only=True):
        joined = " ".join(str(c).strip().lower() if c else "" for c in row)
        if "ncm" in joined:
            wb.close(); file.seek(0); return "fiscal"
        if "custo final" in joined:
            wb.close(); file.seek(0); return "fisico"
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 7:
            continue
        if is_data_row(row):
            col1 = str(row[1]).strip() if row[1] else ""
            wb.close(); file.seek(0)
            return "fiscal" if col1.isdigit() and len(col1) == 8 else "fisico"
    wb.close(); file.seek(0)
    return "fisico"


def process_cplus_file(file, col_map: dict) -> pd.DataFrame:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(values_only=True):
        if is_data_row(row):
            records.append({cn: row[ci] if ci < len(row) else None for cn, ci in col_map.items()})
    wb.close()
    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["Código"] = df["Código"].astype(str).str.strip()
    for col in ["Produto", "NCM"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace({"None": "", "nan": ""})
    if "UN" in df.columns:
        df["UN"] = df["UN"].fillna("").astype(str).str.strip().replace({"None": "", "nan": ""})
    for col in [c for c in df.columns if c not in ("Código", "Produto", "NCM", "UN")]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def smart_load(file):
    layout = detect_layout(file)
    col_map = COL_MAP_FISCAL if layout == "fiscal" else COL_MAP_FISICO
    return process_cplus_file(file, col_map), layout


def to_excel(dfs: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            if df.empty:
                continue
            name = sheet[:31]
            df.to_excel(writer, index=False, sheet_name=name)
            ws = writer.sheets[name]
            hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            for cell in ws[1]:
                cell.fill = hf
                cell.font = Font(color="FFFFFF", bold=True, size=11)
                cell.alignment = Alignment(horizontal="center")
            for col_cells in ws.columns:
                mx = max(len(str(c.value or "")) for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(mx + 4, 60)
    return output.getvalue()


def render_metrics_bar(df, label_qtde, label_valor, col_qtde, col_valor):
    """Exibe linha de métricas resumo para uma aba."""
    c1, c2, c3 = st.columns(3)
    c1.metric("Total de Itens", f"{len(df):,}")
    c2.metric(label_qtde, f"{df[col_qtde].sum():,.0f}")
    c3.metric(label_valor, f"R$ {df[col_valor].sum():,.2f}")


def render_search(df, key):
    """Campo de busca reutilizável."""
    busca = st.text_input("🔍 Buscar por código ou produto", key=key)
    if busca:
        mask = df["Código"].str.contains(busca, case=False, na=False)
        if "Produto" in df.columns:
            mask = mask | df["Produto"].astype(str).str.contains(busca, case=False, na=False)
        df = df[mask]
    return df


# ══════════════════════════════════════════════════════════════════════════════
# INTERFACE
# ══════════════════════════════════════════════════════════════════════════════

st.title("📦 Gestão de Estoque — Físico x Fiscal")

tab_cons, tab_comp = st.tabs(["📋 Consolidar Planilha", "🔍 Comparar Físico x Fiscal"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB — CONSOLIDADOR
# ══════════════════════════════════════════════════════════════════════════════
with tab_cons:
    st.subheader("Consolidar planilha de inventário")
    st.markdown("Envie o arquivo bruto do C-Plus (físico ou fiscal). O layout é detectado automaticamente.")

    file_cons = st.file_uploader("📁 Arquivo (.xlsx)", type=["xlsx"], key="consolidar")
    if file_cons:
        with st.spinner("Processando..."):
            df_c, lay_c = smart_load(file_cons)
        tipo = lay_c.upper()
        st.success(f"✅ Layout: **{tipo}** — **{len(df_c):,}** produtos")
        render_metrics_bar(df_c, "Qtde Total", "Valor Total", "Qtde", "Valor Total")
        st.divider()
        df_show = render_search(df_c, "busca_c")
        st.dataframe(df_show, use_container_width=True, hide_index=True, height=500)
        st.download_button(
            "⬇️ Baixar consolidado (.xlsx)",
            to_excel({"Consolidado": df_show}),
            file_name=f"estoque_{tipo.lower()}_consolidado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

# ══════════════════════════════════════════════════════════════════════════════
# TAB — COMPARAÇÃO
# ══════════════════════════════════════════════════════════════════════════════
with tab_comp:
    st.subheader("Comparar estoque Físico x Fiscal")

    cu1, cu2 = st.columns(2)
    with cu1:
        f_fis = st.file_uploader("📦 Estoque FÍSICO (.xlsx)", type=["xlsx"], key="fisico")
    with cu2:
        f_fisc = st.file_uploader("📑 Estoque FISCAL (.xlsx)", type=["xlsx"], key="fiscal")

    if f_fis and f_fisc:
        with st.spinner("Consolidando e comparando..."):
            df_fis, l_fis = smart_load(f_fis)
            df_fisc, l_fisc = smart_load(f_fisc)

        st.info(
            f"📦 Físico: **{len(df_fis):,}** itens  |  "
            f"📑 Fiscal: **{len(df_fisc):,}** itens"
        )

        # ── Preparar merge ─────────────────────────────────────────────
        fis = df_fis[["Código", "Produto", "Qtde", "Valor Total"]].copy()
        fis.columns = ["Código", "Produto_Fís", "Qtde_Físico", "VT_Físico"]

        cols_fisc = ["Código", "Produto", "Qtde", "Valor Total"]
        if "NCM" in df_fisc.columns:
            cols_fisc.insert(1, "NCM")
        fisc = df_fisc[cols_fisc].copy()
        fisc = fisc.rename(columns={"Produto": "Produto_Fisc", "Qtde": "Qtde_Fiscal", "Valor Total": "VT_Fiscal"})

        merged = pd.merge(fis, fisc, on="Código", how="outer", indicator=True)
        for c in ["Qtde_Físico", "Qtde_Fiscal", "VT_Físico", "VT_Fiscal"]:
            merged[c] = merged[c].fillna(0)
        merged["Produto"] = merged["Produto_Fís"].fillna("").replace({"": None})
        merged["Produto"] = merged["Produto"].fillna(merged.get("Produto_Fisc", "")).fillna("")

        # ── 1) SÓ NO FÍSICO ───────────────────────────────────────────
        so_fisico = merged[merged["_merge"] == "left_only"].copy()
        so_fisico = so_fisico[["Código", "Produto", "Qtde_Físico", "VT_Físico"]]
        so_fisico.columns = ["Código", "Produto", "Qtde", "Valor Total"]
        so_fisico = so_fisico.sort_values("Valor Total", ascending=False)

        # ── 2) SÓ NO FISCAL ───────────────────────────────────────────
        so_fiscal = merged[merged["_merge"] == "right_only"].copy()
        cols_sf = ["Código", "Produto_Fisc", "Qtde_Fiscal", "VT_Fiscal"]
        if "NCM" in merged.columns:
            cols_sf.insert(1, "NCM")
        so_fiscal = so_fiscal[cols_sf]
        rename_sf = {"Produto_Fisc": "Produto", "Qtde_Fiscal": "Qtde", "VT_Fiscal": "Valor Total"}
        so_fiscal = so_fiscal.rename(columns=rename_sf)
        so_fiscal = so_fiscal.sort_values("Valor Total", ascending=False)

        # ── 3) EM AMBOS — com divergências explicadas ─────────────────
        ambos = merged[merged["_merge"] == "both"].copy()
        ambos["Dif_Qtde"] = ambos["Qtde_Físico"] - ambos["Qtde_Fiscal"]
        ambos["Dif_Valor"] = ambos["VT_Físico"] - ambos["VT_Fiscal"]

        def gerar_analise(row):
            dq = row["Dif_Qtde"]
            qf = int(row["Qtde_Físico"])
            qc = int(row["Qtde_Fiscal"])
            if dq == 0:
                return "✅ Quantidades iguais"
            elif dq > 0:
                return f"⚠️ Físico tem {int(dq)} a MAIS que o Fiscal ({qf} vs {qc})"
            else:
                return f"⚠️ Fiscal tem {int(abs(dq))} a MAIS que o Físico ({qf} vs {qc})"

        ambos["Análise"] = ambos.apply(gerar_analise, axis=1)

        cols_ambos = ["Código", "Produto"]
        if "NCM" in ambos.columns:
            cols_ambos.append("NCM")
        cols_ambos += ["Qtde_Físico", "Qtde_Fiscal", "Dif_Qtde",
                       "VT_Físico", "VT_Fiscal", "Dif_Valor", "Análise"]
        ambos = ambos[cols_ambos]
        ambos = ambos.sort_values("Dif_Qtde", key=abs, ascending=False)

        divergentes = ambos[ambos["Dif_Qtde"] != 0]
        iguais = ambos[ambos["Dif_Qtde"] == 0]

        # ══════════════════════════════════════════════════════════════
        # RESUMO GERAL (antes das sub-abas)
        # ══════════════════════════════════════════════════════════════
        st.divider()
        st.markdown("### 📊 Panorama Geral")

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Só no Físico", f"{len(so_fisico):,}")
        m2.metric("Só no Fiscal", f"{len(so_fiscal):,}")
        m3.metric("Em ambos (divergentes)", f"{len(divergentes):,}")
        m4.metric("Em ambos (iguais)", f"{len(iguais):,}")

        v1, v2, v3 = st.columns(3)
        v1.metric("💰 Valor só no Físico", f"R$ {so_fisico['Valor Total'].sum():,.2f}")
        v2.metric("💰 Valor só no Fiscal", f"R$ {so_fiscal['Valor Total'].sum():,.2f}")
        dif_ambos = divergentes["Dif_Valor"].sum() if not divergentes.empty else 0
        v3.metric("💰 Dif. nos que estão em ambos", f"R$ {dif_ambos:,.2f}")

        # ══════════════════════════════════════════════════════════════
        # SUB-ABAS DA COMPARAÇÃO
        # ══════════════════════════════════════════════════════════════
        st.divider()

        sub1, sub2, sub3 = st.tabs([
            f"📦 Só no Físico ({len(so_fisico):,})",
            f"📑 Só no Fiscal ({len(so_fiscal):,})",
            f"🔄 Em Ambos — Divergências ({len(ambos):,})",
        ])

        # ── Sub-aba 1: Só no Físico ────────────────────────────────────
        with sub1:
            st.markdown(
                "**Itens que existem no estoque físico mas NÃO constam no fiscal.** "
                "Pode indicar entrada sem nota, doação, bonificação ou erro de cadastro."
            )
            render_metrics_bar(so_fisico, "Qtde Total", "Valor Total", "Qtde", "Valor Total")
            st.divider()
            df_v1 = render_search(so_fisico, "b_sf")
            st.markdown(f"**{len(df_v1):,} registros**")
            st.dataframe(
                df_v1, use_container_width=True, hide_index=True, height=500,
                column_config={
                    "Qtde": st.column_config.NumberColumn("Qtde", format="%d"),
                    "Valor Total": st.column_config.NumberColumn("Valor Total", format="R$ %.2f"),
                },
            )

        # ── Sub-aba 2: Só no Fiscal ────────────────────────────────────
        with sub2:
            st.markdown(
                "**Itens que constam no fiscal mas NÃO foram encontrados no estoque físico.** "
                "Pode indicar venda sem baixa, perda, furto ou divergência de cadastro."
            )
            render_metrics_bar(so_fiscal, "Qtde Total", "Valor Total", "Qtde", "Valor Total")
            st.divider()
            df_v2 = render_search(so_fiscal, "b_sfisc")
            st.markdown(f"**{len(df_v2):,} registros**")
            st.dataframe(
                df_v2, use_container_width=True, hide_index=True, height=500,
                column_config={
                    "Qtde": st.column_config.NumberColumn("Qtde", format="%d"),
                    "Valor Total": st.column_config.NumberColumn("Valor Total", format="R$ %.2f"),
                },
            )

        # ── Sub-aba 3: Em Ambos ────────────────────────────────────────
        with sub3:
            st.markdown(
                "**Itens presentes nos dois estoques.** "
                "A coluna **Análise** explica a divergência de cada item."
            )

            # Filtro de situação
            cf1, cf2 = st.columns([2, 1])
            with cf2:
                filtro_sit = st.selectbox("Filtrar por", [
                    "Todos",
                    "Só divergentes",
                    "Só iguais",
                    "Físico tem mais",
                    "Fiscal tem mais",
                ], key="filtro_ambos")

            with cf1:
                busca_ambos = st.text_input("🔍 Buscar por código ou produto", key="b_ambos")

            df_v3 = ambos.copy()

            if filtro_sit == "Só divergentes":
                df_v3 = df_v3[df_v3["Dif_Qtde"] != 0]
            elif filtro_sit == "Só iguais":
                df_v3 = df_v3[df_v3["Dif_Qtde"] == 0]
            elif filtro_sit == "Físico tem mais":
                df_v3 = df_v3[df_v3["Dif_Qtde"] > 0]
            elif filtro_sit == "Fiscal tem mais":
                df_v3 = df_v3[df_v3["Dif_Qtde"] < 0]

            if busca_ambos:
                mask = (
                    df_v3["Código"].str.contains(busca_ambos, case=False, na=False)
                    | df_v3["Produto"].astype(str).str.contains(busca_ambos, case=False, na=False)
                )
                df_v3 = df_v3[mask]

            # Métricas da seleção atual
            mc1, mc2, mc3, mc4 = st.columns(4)
            mc1.metric("Exibindo", f"{len(df_v3):,} itens")
            n_div = (df_v3["Dif_Qtde"] != 0).sum()
            n_ok = (df_v3["Dif_Qtde"] == 0).sum()
            mc2.metric("Divergentes", f"{n_div:,}")
            mc3.metric("Iguais", f"{n_ok:,}")
            mc4.metric("Dif. Valor Total", f"R$ {df_v3['Dif_Valor'].sum():,.2f}")

            st.dataframe(
                df_v3, use_container_width=True, hide_index=True, height=600,
                column_config={
                    "Código": st.column_config.TextColumn("Código", width="small"),
                    "Produto": st.column_config.TextColumn("Produto", width="large"),
                    "NCM": st.column_config.TextColumn("NCM", width="small"),
                    "Qtde_Físico": st.column_config.NumberColumn("Qtde Físico", format="%d"),
                    "Qtde_Fiscal": st.column_config.NumberColumn("Qtde Fiscal", format="%d"),
                    "Dif_Qtde": st.column_config.NumberColumn("Dif. Qtde", format="%d"),
                    "VT_Físico": st.column_config.NumberColumn("VT Físico", format="R$ %.2f"),
                    "VT_Fiscal": st.column_config.NumberColumn("VT Fiscal", format="R$ %.2f"),
                    "Dif_Valor": st.column_config.NumberColumn("Dif. Valor", format="R$ %.2f"),
                    "Análise": st.column_config.TextColumn("Análise", width="large"),
                },
            )

        # ══════════════════════════════════════════════════════════════
        # DOWNLOAD GERAL
        # ══════════════════════════════════════════════════════════════
        st.divider()
        sheets = {
            "Só no Físico": so_fisico,
            "Só no Fiscal": so_fiscal,
            "Ambos - Comparativo": ambos,
        }
        st.download_button(
            "⬇️ Baixar relatório completo (.xlsx)",
            to_excel(sheets),
            file_name="comparativo_fisico_x_fiscal.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
